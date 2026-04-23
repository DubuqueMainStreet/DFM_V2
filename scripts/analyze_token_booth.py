"""
Token booth: card transactions (Excel) + weekly zip logs (OCR from scanned PDFs).

Zip rows: full-page RapidOCR for 5-digit zips. Market dates: OCR on the top-left crop
of each page (where the worksheet prints "Date:"), with forward-fill within each PDF.
If a page still has no date, fall back to imputation from PDF_SATURDAY_RANGES + card
transaction weights (see impute_market_dates_from_pdf_pages).

Outputs: scripts/output/zip_observations.csv, token_booth_weekly_metrics.csv,
token_booth_reconciliation_by_day.csv, token_booth_event_week_summary.txt,
token_booth_weekly_Neff.png

Default inputs: S:\\ (override with env TOKEN_BOOTH_ROOT or --root).

Usage:
  python scripts/analyze_token_booth.py --extract-pdfs   # OCR + write zip CSV + metrics
  python scripts/analyze_token_booth.py                  # metrics from cached zip CSV
"""
from __future__ import annotations

import argparse
import csv
import math
import os
import re
import sys
from collections import Counter, defaultdict
import bisect
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Iterator, Optional

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = REPO_ROOT / "scripts" / "output"
DEFAULT_ROOT = os.environ.get(
    "TOKEN_BOOTH_ROOT",
    r"s:\FARMERS MARKET\2025 - 180th\Tokens and Market Money",
)

ZIP_CSV = OUTPUT_DIR / "zip_observations.csv"
WEEKLY_METRICS_CSV = OUTPUT_DIR / "token_booth_weekly_metrics.csv"
CARD_DAILY_CSV = OUTPUT_DIR / "token_booth_card_by_market_day.csv"
RECONCILE_CSV = OUTPUT_DIR / "token_booth_reconciliation_by_day.csv"
EVENTS_CSV = REPO_ROOT / "scripts" / "data" / "dubuque_events_2025.csv"
SUMMARY_TXT = OUTPUT_DIR / "token_booth_event_week_summary.txt"
CHART_PNG = OUTPUT_DIR / "token_booth_weekly_Neff.png"

WEEKLY_PDFS = [
    "May and June 2025 weekly transaction work sheets.pdf",
    "July through Sept 2025 weekly transaction work sheets.pdf",
    "Oct 2025 weekly transaction log.pdf",
]

XLSX_NAME = "2025 Transactions at DFM.xlsx"

# Max PDF page index gap after a successful top-left date OCR to still reuse that date
# (continuation sheets in the same week). Beyond this, leave date empty for imputation.
DATE_FORWARD_FILL_MAX_GAP = 4

# Saturday market days covered by each scanned workbook (inclusive ranges).
PDF_SATURDAY_RANGES: dict[str, tuple[date, date]] = {
    "May and June 2025 weekly transaction work sheets.pdf": (date(2025, 5, 3), date(2025, 6, 28)),
    "July through Sept 2025 weekly transaction work sheets.pdf": (date(2025, 7, 5), date(2025, 9, 27)),
    "Oct 2025 weekly transaction log.pdf": (date(2025, 10, 4), date(2025, 10, 25)),
}


def saturdays_between(start: date, end: date) -> list[date]:
    out: list[date] = []
    d = start
    while d <= end:
        if d.weekday() == 5:
            out.append(d)
        d += timedelta(days=1)
    return out


def impute_market_dates_from_pdf_pages(
    rows: list[dict],
    card_by_day: dict[date, dict],
    only_when_missing: bool = True,
) -> list[dict]:
    """
    Assign market_date from page position + Saturday list + card-tx weights.
    Used only when top-left date OCR did not yield a date (if only_when_missing).
    """
    for pdf, (start, end) in PDF_SATURDAY_RANGES.items():
        all_for_pdf = [r for r in rows if r.get("source_pdf") == pdf]
        if not all_for_pdf:
            continue
        chunk = all_for_pdf
        if only_when_missing:
            chunk = [r for r in chunk if not (r.get("market_date") or "").strip()]
        if not chunk:
            continue
        sats = saturdays_between(start, end)
        if not sats:
            continue
        weights = [float(card_by_day.get(d, {}).get("tx_count", 0) or 0) for d in sats]
        tw = sum(weights)
        if tw <= 0:
            weights = [1.0] * len(sats)
            tw = sum(weights)
        cumw = [0.0]
        for w in weights:
            cumw.append(cumw[-1] + w)
        n_pages = max(int(r.get("page") or 0) for r in all_for_pdf)
        n_pages = max(n_pages, 1)
        for r in chunk:
            p = int(r.get("page") or 0)
            frac = (p - 0.5) / n_pages
            frac = min(1.0, max(0.0, frac))
            target = frac * tw
            idx = bisect.bisect_right(cumw, target) - 1
            idx = max(0, min(len(sats) - 1, idx))
            r["market_date"] = sats[idx].isoformat()
            r["date_source"] = "imputed"
    return rows


# -----------------------------------------------------------------------------
# Zip normalization
# -----------------------------------------------------------------------------

def is_plausible_zip(z: str) -> bool:
    if len(z) != 5 or not z.isdigit():
        return False
    n = int(z)
    if n < 501 or n > 99950:  # rough US range; excludes 00000
        return False
    return True


def normalize_zip_token(raw: str) -> Optional[str]:
    """Normalize OCR token to 5-digit zip or None."""
    t = raw.strip().replace(" ", "").replace("O", "0").replace("o", "0")
    # Trailing letter glitch e.g. 52001g
    t = re.sub(r"^(\d{5})[a-zA-Z]$", r"\1", t)
    if not t:
        return None
    # Common 6-char glitch: leading 1 before 520xx
    if len(t) == 6 and t[0] == "1" and t[1:3] == "52" and t[2:].isdigit():
        t = t[1:]
    if len(t) == 5 and t.isdigit() and is_plausible_zip(t):
        return t
    # 3-digit local shorthand -> 52xxx (Dubuque area)
    if len(t) == 3 and t.isdigit():
        n = int(t)
        if n <= 99:
            cand = f"520{n:02d}"
            if is_plausible_zip(cand):
                return cand
    return None


# -----------------------------------------------------------------------------
# Top-left date OCR (worksheet prints "Date:" in the top-left corner)
# -----------------------------------------------------------------------------

DATE_LABEL_RE = re.compile(r"Date:\s*(\d{1,2})/(\d{1,2})/(\d{2,4})", re.I)
DATE_ANY_RE = re.compile(r"\b(\d{1,2})/(\d{1,2})/(\d{2,4})\b")
# Handwritten / scan: "5/25/25" misread as "0-25- 25" (lost leading month digit)
DATE_DASH_LOOSE_RE = re.compile(
    r"\b(\d{1,2})\s*[-]\s*(\d{1,2})\s*[-\s]+\s*(\d{2,4})\b"
)


def parse_mdy_to_date(m: str, d: str, y: str) -> Optional[date]:
    try:
        yi = int(y)
        if yi < 100:
            yi += 2000
        return date(yi, int(m), int(d))
    except (ValueError, TypeError):
        return None


def normalize_ocr_market_date(d: Optional[date]) -> Optional[date]:
    """Fix common OCR year/month confusions; keep 2025 DFM season (May–Oct)."""
    if d is None:
        return None
    y, mo, day = d.year, d.month, d.day
    if y in (2023, 2024) and 5 <= mo <= 10:
        y = 2025
    if y == 2025 and mo == 2:
        mo = 5
    try:
        d2 = date(y, mo, day)
    except ValueError:
        return None
    if d2 < date(2025, 5, 1) or d2 > date(2025, 10, 31):
        return None
    return d2


def month_try_order_for_pdf(pdf_basename: str) -> tuple[int, ...]:
    """Narrow ambiguous month when OCR drops the leading digit (e.g. 0-25-25)."""
    if "May and June" in pdf_basename:
        return (5, 6)
    if "July through Sept" in pdf_basename:
        return (7, 8, 9)
    if "Oct" in pdf_basename:
        return (10,)
    return (5, 6, 7, 8, 9, 10)


def parse_dash_date_loose(full: str, month_order: tuple[int, ...]) -> Optional[date]:
    """Parse dashed tokens like '0-25- 25' when the month digit is garbled (often reads as 0)."""
    m = DATE_DASH_LOOSE_RE.search(full)
    if not m:
        return None
    mi, di, yi = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if yi < 100:
        yi += 2000
    if mi == 0 and 1 <= di <= 31:
        for month in month_order:
            try:
                d = normalize_ocr_market_date(date(yi, month, di))
                if d:
                    return d
            except ValueError:
                continue
        return None
    try:
        return normalize_ocr_market_date(date(yi, mi, di))
    except ValueError:
        return None


def parse_market_date_from_crop_text(full: str, pdf_basename: str) -> Optional[date]:
    """Prefer labeled Date:; else slash dates; else dashed OCR artifacts."""
    mo = month_try_order_for_pdf(pdf_basename)
    m = DATE_LABEL_RE.search(full)
    if m:
        d = parse_mdy_to_date(m.group(1), m.group(2), m.group(3))
        d = normalize_ocr_market_date(d)
        if d:
            return d
    for m2 in DATE_ANY_RE.finditer(full):
        d = parse_mdy_to_date(m2.group(1), m2.group(2), m2.group(3))
        d = normalize_ocr_market_date(d)
        if d:
            return d
    return parse_dash_date_loose(full, mo)


def parse_market_date_from_ocr_boxes(ocr_result: list, pdf_basename: str) -> Optional[date]:
    """Sort OCR boxes top-to-bottom, left-to-right, then parse date string."""
    if not ocr_result:
        return None
    items: list[tuple[float, float, str]] = []
    for item in ocr_result:
        if not item or len(item) < 2:
            continue
        if not isinstance(item[1], str):
            continue
        box = item[0]
        if not box:
            continue
        try:
            y = float(box[0][1])
            x = float(box[0][0])
        except (TypeError, IndexError, ValueError):
            x, y = 0.0, 0.0
        items.append((y, x, item[1].strip()))
    items.sort(key=lambda t: (t[0], t[1]))
    full = " ".join(t[2] for t in items)
    return parse_market_date_from_crop_text(full, pdf_basename)


def ocr_top_left_date_for_page(page, ocr_engine, pdf_basename: str) -> Optional[date]:
    """OCR only the top-left region where the worksheet date is printed."""
    import fitz  # PyMuPDF

    r = page.rect
    clip = fitz.Rect(
        r.x0,
        r.y0,
        r.x0 + r.width * 0.52,
        r.y0 + r.height * 0.34,
    )
    pix = page.get_pixmap(dpi=220, clip=clip)
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        pix.save(tmp_path)
        ocr_out = ocr_engine(tmp_path)
        result = ocr_out[0] if ocr_out else None
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
    return parse_market_date_from_ocr_boxes(result or [], pdf_basename)


# -----------------------------------------------------------------------------
# PDF OCR -> zip rows
# -----------------------------------------------------------------------------

def extract_zips_from_ocr_result(ocr_result: list) -> list[tuple[str, Optional[float]]]:
    """RapidOCR returns list of [box, text, score]."""
    out: list[tuple[str, Optional[float]]] = []
    for item in ocr_result:
        if not item or len(item) < 2:
            continue
        text = item[1]
        if not isinstance(text, str):
            continue
        score = float(item[2]) if len(item) > 2 and item[2] is not None else None
        t = text.strip()
        if not t or t.startswith("$"):
            continue
        if re.search(r"\$\s*\d", t):
            continue
        # single token
        z = normalize_zip_token(t)
        if z:
            out.append((z, score))
            continue
        # embedded in longer string
        for m in re.finditer(r"\b(\d{5})\b", t):
            z2 = normalize_zip_token(m.group(1))
            if z2:
                out.append((z2, score))
    return out


def ocr_pdf_pages(
    pdf_path: Path, ocr_engine
) -> Iterator[tuple[int, list[tuple[str, Optional[float]]], Optional[date]]]:
    import fitz

    pdf_base = pdf_path.name
    doc = fitz.open(str(pdf_path))
    for page_idx in range(doc.page_count):
        page = doc[page_idx]
        d_crop = ocr_top_left_date_for_page(page, ocr_engine, pdf_base)
        pix = page.get_pixmap(dpi=150)
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            pix.save(tmp_path)
            ocr_out = ocr_engine(tmp_path)
            result = ocr_out[0] if ocr_out else None
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        zips = extract_zips_from_ocr_result(result or [])
        yield page_idx, zips, d_crop
    doc.close()


def run_pdf_extraction(root: Path) -> Path:
    try:
        from rapidocr_onnxruntime import RapidOCR
    except ImportError:
        print("Install: pip install rapidocr-onnxruntime onnxruntime pymupdf", file=sys.stderr)
        raise

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ocr_engine = RapidOCR()
    rows: list[dict[str, object]] = []
    for name in WEEKLY_PDFS:
        pdf = root / name
        if not pdf.exists():
            print(f"skip missing: {pdf}", file=sys.stderr)
            continue
        print(f"OCR {name} ...")
        last_ff: Optional[date] = None
        last_ff_page = -100
        for page_idx, zips, d_crop in ocr_pdf_pages(pdf, ocr_engine):
            if d_crop is not None:
                last_ff = d_crop
                last_ff_page = page_idx
                eff: Optional[date] = d_crop
            else:
                if (
                    last_ff is not None
                    and (page_idx - last_ff_page) <= DATE_FORWARD_FILL_MAX_GAP
                ):
                    eff = last_ff
                else:
                    eff = None
            for z, sc in zips:
                rows.append(
                    {
                        "source_pdf": name,
                        "page": page_idx + 1,
                        "market_date": eff.isoformat() if eff else "",
                        "date_source": "ocr" if eff else "",
                        "zip": z,
                        "ocr_score": sc if sc is not None else "",
                    }
                )
    xlsx = root / XLSX_NAME
    if xlsx.exists():
        cards = load_card_transactions(xlsx)
        by_day = card_summary_by_day(cards)
        rows = impute_market_dates_from_pdf_pages(rows, by_day, only_when_missing=True)
    with ZIP_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["source_pdf", "page", "market_date", "date_source", "zip", "ocr_score"],
        )
        w.writeheader()
        w.writerows(rows)
    print(f"Wrote {len(rows)} zip rows -> {ZIP_CSV}")
    return ZIP_CSV


# -----------------------------------------------------------------------------
# Excel card transactions
# -----------------------------------------------------------------------------

HEADER_ROW = 0


def parse_excel_datetime(cell) -> Optional[datetime]:
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return cell
    s = str(cell).strip()
    if not s:
        return None
    # Strip trailing timezone tokens (CDT, CST, etc.)
    s = re.sub(r"\s+(CDT|CST|EST|PST|EDT|UTC)\s*$", "", s, flags=re.I)
    for fmt in ("%d-%b-%Y %I:%M %p", "%d-%b-%Y %H:%M:%S"):
        try:
            return datetime.strptime(s[:30], fmt)
        except ValueError:
            continue
    return None


def load_card_transactions(xlsx_path: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out: list[dict] = []
    payment_sheets = [n for n in wb.sheetnames if n.startswith("Payments-")]
    sheet_order = payment_sheets + ["Jazz Sales", "EBT Transactions"]
    for sheet_name in sheet_order:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        idx_date = idx_result = idx_amount = idx_tender = None
        if header and any(h is not None and str(h).strip() for h in header):
            for i, h in enumerate(header):
                if not h:
                    continue
                hs = str(h).strip().lower()
                if hs == "payment date":
                    idx_date = i
                elif hs == "result":
                    idx_result = i
                elif hs == "amount":
                    idx_amount = i
                elif hs == "tender":
                    idx_tender = i
            data_rows = rows[1:]
        else:
            # Jazz / EBT: no header row; fixed columns match Payments export
            idx_date, idx_amount, idx_result, idx_tender = 0, 5, 8, 1
            data_rows = rows
        for row in data_rows:
            if not row or idx_date is None:
                continue
            if idx_result is not None and idx_result < len(row) and row[idx_result] is not None:
                if str(row[idx_result]).strip().upper() != "SUCCESS":
                    continue
            dt = parse_excel_datetime(row[idx_date] if idx_date < len(row) else None)
            if not dt:
                continue
            amt = row[idx_amount] if idx_amount is not None and idx_amount < len(row) else None
            tender = row[idx_tender] if idx_tender is not None and idx_tender < len(row) else ""
            try:
                amt_f = float(amt) if amt is not None else 0.0
            except (TypeError, ValueError):
                amt_f = 0.0
            out.append(
                {
                    "sheet": sheet_name,
                    "dt": dt,
                    "market_day": dt.date(),
                    "amount": amt_f,
                    "tender": tender or "",
                }
            )
    wb.close()
    return out


def card_summary_by_day(cards: list[dict]) -> dict[date, dict]:
    by: dict[date, dict] = defaultdict(lambda: {"tx_count": 0, "revenue": 0.0})
    for c in cards:
        d = c["market_day"]
        by[d]["tx_count"] += 1
        try:
            by[d]["revenue"] += float(c["amount"])
        except (TypeError, ValueError):
            pass
    return dict(by)


def week_key_saturday_anchor(d: date) -> tuple[int, int]:
    """ISO year/week where week is anchored by the Saturday market day (week containing d)."""
    # Shift so that Saturday is end of "market week": use ISO week of (d + 2 days) or standard ISO
    return d.isocalendar()[:2]


# -----------------------------------------------------------------------------
# Diversity metrics
# -----------------------------------------------------------------------------

def shannon_entropy(counts: Counter) -> float:
    total = sum(counts.values())
    if total <= 0:
        return 0.0
    h = 0.0
    for c in counts.values():
        if c <= 0:
            continue
        p = c / total
        h -= p * math.log(p)
    return h


def diversity_metrics(counts: Counter) -> dict:
    total = sum(counts.values())
    k = len(counts)
    h = shannon_entropy(counts)
    n_eff = math.exp(h) if total > 0 else 0.0
    simpson = 0.0
    if total > 0:
        simpson = 1.0 - sum((c / total) ** 2 for c in counts.values())
    return {
        "n_observations": total,
        "unique_zips": k,
        "shannon_H": round(h, 4),
        "N_eff": round(n_eff, 4),
        "simpson": round(simpson, 4),
    }


def pct_outside_520xx_for_week(zip_rows: list[dict]) -> dict[tuple[int, int], float]:
    """Share of zip observations not starting with 520 (rough non-local / non-core-Dubuque proxy)."""
    by_week: dict[tuple[int, int], list[int]] = defaultdict(lambda: [0, 0])
    for r in zip_rows:
        md = r.get("market_date") or ""
        if not md:
            continue
        try:
            d = date.fromisoformat(md)
        except ValueError:
            continue
        z = r.get("zip") or ""
        if len(z) != 5:
            continue
        wk = week_key_saturday_anchor(d)
        by_week[wk][0] += 1
        if not z.startswith("520"):
            by_week[wk][1] += 1
    out_pct: dict[tuple[int, int], float] = {}
    for wk, pair in by_week.items():
        tot, o = pair[0], pair[1]
        out_pct[wk] = round(100.0 * o / tot, 2) if tot else 0.0
    return out_pct


def aggregate_zips_by_week(zip_rows: list[dict]) -> dict[tuple[int, int], Counter]:
    """Week -> zip -> count."""
    by_week: dict[tuple[int, int], Counter] = defaultdict(Counter)
    for r in zip_rows:
        md = r.get("market_date") or ""
        if not md:
            continue
        try:
            d = date.fromisoformat(md)
        except ValueError:
            continue
        z = r.get("zip") or ""
        if not z:
            continue
        wk = week_key_saturday_anchor(d)
        by_week[wk][z] += 1
    return dict(by_week)


def load_zip_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_events(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def event_week_flags(
    events: list[dict], weeks: Iterable[tuple[int, int]]
) -> dict[tuple[int, int], list[str]]:
    """Map week -> list of event names overlapping that ISO week."""
    from datetime import timedelta

    flags: dict[tuple[int, int], list[str]] = defaultdict(list)
    week_set = set(weeks)
    for ev in events:
        try:
            s = date.fromisoformat(ev["start_date"].strip())
            e = date.fromisoformat(ev["end_date"].strip())
        except (KeyError, ValueError):
            continue
        d = s
        while d <= e:
            wk = week_key_saturday_anchor(d)
            if wk in week_set and ev.get("event_name"):
                name = ev["event_name"].strip()
                if name and name not in flags[wk]:
                    flags[wk].append(name)
            d += timedelta(days=1)
    return dict(flags)


def plot_weekly(
    weekly_rows: list[dict],
    event_flags: dict[tuple[int, int], list[str]],
    out_png: Path,
) -> None:
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
    except ImportError:
        print("matplotlib not installed; skip chart", file=sys.stderr)
        return

    weekly_rows = sorted(weekly_rows, key=lambda r: (r["iso_year"], r["iso_week"]))
    xs = []
    ys = []
    for r in weekly_rows:
        iy, iw = int(r["iso_year"]), int(r["iso_week"])
        # Thursday of ISO week for x
        d = datetime.fromisocalendar(iy, iw, 4).date()
        xs.append(d)
        ys.append(float(r["N_eff"]))
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.plot(xs, ys, marker="o", linewidth=1.5, label="N_eff (zip diversity)")
    # shade weeks that have events
    for r in weekly_rows:
        iy, iw = int(r["iso_year"]), int(r["iso_week"])
        wk = (iy, iw)
        if event_flags.get(wk):
            d0 = datetime.fromisocalendar(iy, iw, 1).date()
            d1 = datetime.fromisocalendar(iy, iw, 7).date()
            ax.axvspan(d0, d1, alpha=0.15, color="orange")
    ax.set_title("Weekly zip diversity (effective number of zips) — shaded weeks have listed events")
    ax.set_ylabel("N_eff")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
    ax.legend()
    fig.autofmt_xdate()
    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=120)
    plt.close(fig)
    print(f"Wrote {out_png}")


def write_reconciliation(zip_rows: list[dict], by_day: dict[date, dict]) -> None:
    zip_counts: dict[date, int] = defaultdict(int)
    for r in zip_rows:
        md = r.get("market_date") or ""
        if not md:
            continue
        try:
            d = date.fromisoformat(md)
        except ValueError:
            continue
        zip_counts[d] += 1
    all_days = sorted(set(by_day.keys()) | set(zip_counts.keys()))
    RECONCILE_CSV.parent.mkdir(parents=True, exist_ok=True)
    with RECONCILE_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["market_day", "card_tx_count", "zip_rows_ocr", "delta_zip_minus_card"])
        for d in all_days:
            c = by_day.get(d, {}).get("tx_count", 0)
            z = zip_counts.get(d, 0)
            w.writerow([d.isoformat(), c, z, z - c])
    print(f"Wrote {RECONCILE_CSV}")


def write_event_summary(
    weekly_rows: list[dict],
    event_flags: dict[tuple[int, int], list[str]],
    out_path: Path,
) -> None:
    with_event: list[float] = []
    without: list[float] = []
    for r in weekly_rows:
        iy, iw = int(r["iso_year"]), int(r["iso_week"])
        ne = float(r["N_eff"])
        if event_flags.get((iy, iw)):
            with_event.append(ne)
        else:
            without.append(ne)
    def avg(a: list[float]) -> float:
        return sum(a) / len(a) if a else 0.0

    lines = [
        "Token booth — event week vs baseline (descriptive only; not causal)",
        f"Weeks with any listed event: n={len(with_event)}, mean N_eff={avg(with_event):.2f}",
        f"Weeks with no listed event: n={len(without)}, mean N_eff={avg(without):.2f}",
        "",
        "Per-week N_eff and overlapping events:",
    ]
    for r in sorted(weekly_rows, key=lambda x: (x["iso_year"], x["iso_week"])):
        wk = (int(r["iso_year"]), int(r["iso_week"]))
        evs = event_flags.get(wk, [])
        lines.append(
            f"  {r['iso_year']}-W{r['iso_week']:02d}: N_eff={r['N_eff']}, unique={r['unique_zips']}, n_zip_obs={r['n_observations']}"
            + (f" | events: {', '.join(evs)}" if evs else "")
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out_path}")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default=DEFAULT_ROOT, help="Folder with xlsx and PDFs")
    ap.add_argument("--extract-pdfs", action="store_true", help="Run OCR and write zip_observations.csv")
    args = ap.parse_args()
    root = Path(args.root)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (REPO_ROOT / "scripts" / "data").mkdir(parents=True, exist_ok=True)

    if args.extract_pdfs:
        run_pdf_extraction(root)

    if not ZIP_CSV.exists():
        print(f"No {ZIP_CSV}; run with --extract-pdfs first.", file=sys.stderr)
        return 1

    xlsx = root / XLSX_NAME
    if not xlsx.exists():
        print(f"Missing Excel: {xlsx}", file=sys.stderr)
        return 1

    cards = load_card_transactions(xlsx)
    by_day = card_summary_by_day(cards)
    with CARD_DAILY_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["market_day", "tx_count", "revenue"])
        for d in sorted(by_day.keys()):
            w.writerow([d.isoformat(), by_day[d]["tx_count"], f"{by_day[d]['revenue']:.2f}"])

    zip_rows = load_zip_csv(ZIP_CSV)
    if not zip_rows:
        print("No zip observations loaded.", file=sys.stderr)
        return 1

    zip_rows = impute_market_dates_from_pdf_pages(zip_rows, by_day, only_when_missing=True)

    write_reconciliation(zip_rows, by_day)

    by_week_counts = aggregate_zips_by_week(zip_rows)
    pct_out = pct_outside_520xx_for_week(zip_rows)
    events = load_events(EVENTS_CSV)
    all_weeks = sorted(by_week_counts.keys())
    event_flags = event_week_flags(events, all_weeks)

    weekly_rows: list[dict] = []
    for wk in all_weeks:
        counts = by_week_counts[wk]
        m = diversity_metrics(counts)
        iy, iw = wk
        row = {
            "iso_year": iy,
            "iso_week": iw,
            "pct_zip_obs_outside_520xx": pct_out.get(wk, 0.0),
            **m,
        }
        weekly_rows.append(row)

    with WEEKLY_METRICS_CSV.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "iso_year",
            "iso_week",
            "pct_zip_obs_outside_520xx",
            "n_observations",
            "unique_zips",
            "shannon_H",
            "N_eff",
            "simpson",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(weekly_rows)

    write_event_summary(weekly_rows, event_flags, SUMMARY_TXT)
    plot_weekly(weekly_rows, event_flags, CHART_PNG)

    print(f"Wrote {WEEKLY_METRICS_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
